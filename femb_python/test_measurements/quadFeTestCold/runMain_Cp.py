from femb_python.test_measurements.quadFeTestCold.scripts.sbnd_femb_meas import FEMB_DAQ
from femb_python.test_measurements.quadFeTestCold.scripts.femb_udp_cmdline import FEMB_UDP
from femb_python.configuration import CONFIG
import os
import sys
#import gc
#import struct
#import copy
#from datetime import datetime
import pickle
import time
import shutil
import time                       #updated
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
from femb_python.test_measurements.quadFeTestCold.user_settings import user_editable_settings
settings = user_editable_settings()
from femb_python.test_measurements.quadFeTestCold.scripts.Data_Analysis import Data_Analysis
from femb_python.test_measurements.quadFeTestCold.scripts.Data_Analysis2 import Data_Analysis2
from femb_python.test_measurements.quadFeTestCold.scripts.Data_Analysis3 import Data_Analysis3

class main_class(object):

    def __init__(self):
        self.sbnd = FEMB_DAQ()
        self.femb = FEMB_UDP()
        self.config = CONFIG()
        self.FunctionGenerator = None
        self.analyze = Data_Analysis()
        self.analyze2 = Data_Analysis2()
        self.analyze3 = Data_Analysis3()

    def check_setup(self):
        self.femb.init_ports(hostIP = settings.PC_IP, destIP = settings.FPGA_IP)
        time.sleep(1)
        self.config.resetFEMBBoard()
        self.config.initBoard()

    def quad_test(self,**params):

#Create main directory for each chip.  Also creates an easily accessible tuple that maps which chips are in which socket.  There are 4 sockets, so something like
# [[0, 'first], [1, 'second'], [2, 'third'], [3, 'fourth']], where the first part is which socket it is and the second is what the user named that chip
#But I also figured that it's very possible that for example only socket 2 could not be working.  In that case, you remove "2" from "chips_to_use" in the
#settings file, and it wont ask for the chip, and it'll be taken into account.  That's what all the gymnastics with the for loop and tuples are for. Use prints if you wanna see what's going on.

        folder_path = [[],[],[],[]]
        chip_name = []
        chip_name.append(params["asic0id"])
        chip_name.append(params["asic1id"])
        chip_name.append(params["asic2id"])
        chip_name.append(params["asic3id"])

        i = 0
        for i in range(0,4):
            now = time.time()
            folder_path[i] = (settings.path+chip_name[i] + time.strftime("_%Y%m%d_T%H%M%S", time.localtime(now)) + "/")
            print("%%%% folder_path: ",folder_path[i])
            try: 
               os.makedirs(folder_path[i])
            except OSError:
               if os.path.exists(folder_path[i]):
                   if (input("Folder already exists.  Overwrite? (y/n)\n") == "y"):
                       for the_file in os.listdir(folder_path[i]):
                           file_path = os.path.join(folder_path[i], the_file)
                           if os.path.isfile(file_path):
                               os.unlink(file_path)
                           else:
                               shutil.rmtree(file_path)

            self.map_directory(folder_path[i])
        print("f1: ", folder_path[0])
        print("f2: ", folder_path[1])
        print("f3: ", folder_path[2])
        print("f4: ", folder_path[3])        

        i = 0
        chip_list = [[],[],[],[]]
        for i in range(settings.chip_num):
            chip_list[i].append(i)
            chip_list[i].append(chip_name[i])
   
        i = 0
        for loops in range(settings.chip_num):
            if (i == (settings.chip_num)):
                break
            if (type(chip_list[i][1]) is list):
                chip_list.pop(i)
                folder_path.pop(i)
            else:
                i = i + 1

        print("Test--> Socket map is {}".format(chip_list))
        
        #Get overall results spreadsheet and start filling it in
        wb = self.get_results_sheet()
        ws = wb.get_sheet_by_name("Results")
        next_space = ws.max_row
        for num,i in enumerate(chip_list):
            ws.cell(row = 1 + next_space + num, column = 1).value = datetime.now().strftime('%Y_%m_%d')
            ws.cell(row = 1 + next_space + num, column = 2).value = i[1]
            ws.cell(row = 1 + next_space + num, column = 3).value = i[0]
            ws.cell(row = 1 + next_space + num, column = 2).alignment = Alignment(horizontal='center')
            ws.cell(row = 1 + next_space + num, column = 3).alignment = Alignment(horizontal='center')
            
        #Create sync directory for chip and sync
        print("Test--> Synchronizing ADCS...")
        chip_indices = []
        for i in range(len(chip_list)):
            chip_indices.append(chip_list[i][0])
        self.sbnd.femb_config.syncADC(chip_indices)
        #Tells the FPGA to turn on each DAC
        self.sbnd.femb_config.femb.write_reg(61, 0x0)
        
        #Read from DATA output ADCs
        self.sbnd.femb_config.femb.write_reg(60, 0)

        #Set to Regular Mode (not sampling scope)
        self.sbnd.femb_config.femb.write_reg(10, 0)
        
        #Select TP FE Mode
        self.sbnd.femb_config.femb.write_reg(9, 3)
        
        #Pulses the internal ASIC pulse (this changes the way packets are outputted through UDP and can screw things up 
        #if you don't know how it works.  Use caution and makes ure register 17 goes back to 0 when done)
        #The way it works is that whenever the FPGA is ready to send out an ASIC pulse, it 
        self.reg_17_value = (settings.default_TP_Period << 16) + (settings.default_TP_Shift << 8) + (0b01000000)
        self.sbnd.femb_config.femb.write_reg(17, self.reg_17_value)
        
        #Print all channels in the chips, so we see what those pulses looked like, so we could go back and see if there was already a problem here
        for num,i in enumerate(chip_list):
            stringy = str(folder_path[num])
            sync_directory = stringy + str("Synchronization/")
            try:
                os.makedirs(sync_directory)
            #except WindowsError:
            except OSError:
                pass
            data = self.sbnd.femb_config.get_data_chipX(chip = i[0], packets = 5)
            print("Test--> Printing synchronization plot for Chip {}".format(i[1]))
            self.sbnd.plot_chips(data = data, plot_name = sync_directory + "Sync_Plot.png", 
                                 title_name = "Pulses for synchronization: Gain = 25 mv/fC, Peaking Time = 3 us, Buffer on, "
                                              "DAC Pulse at {} \nPeaks should be between {} and {}, Baseline should be between "
                                              "{} and {}".format(hex(settings.default_DAC), settings.sync_peak_min, 
                                                settings.sync_peak_max, settings.sync_baseline_min, settings.sync_baseline_max))
                                                
        self.sbnd.femb_config.femb.write_reg(17, 0)
        #A collection of pass/fail messages that will get added to during the process and printed out at the end
        analysis_messages = []

        #Get data for the DAC step tests
        for num,i in enumerate(chip_list):
            self.sbnd.save_DAC_data(folder_path[num], chip_index = i[0], chip_name = i[1])
        #Analyze the DAC step tests
        analysis_messages.append("DAC step Analysis:")
        for num,i in enumerate(chip_list):
            result = self.analyze2.DAC_directory(folder_path[num], i[1])
            cell = ws.cell(row = 1 + next_space + num, column = 6)
            if (result == True):
                analysis_messages.append("Chip {} (Socket {}) PASSED!".format(i[1], i[0]))
                cell.value = "PASS"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
            if (result == False):
                analysis_messages.append("Chip {} (Socket {}) FAILED!".format(i[1], i[0]))
                cell.value = "FAIL"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                
        #Get data for the baseline and RMS tests
        for num,i in enumerate(chip_list):
            self.sbnd.save_rms_noise(folder_path[num], chip_index = i[0], chip_name = i[1])
#Analyze the baseline and RMS tests         
        analysis_messages.append("Baseline and RMS Analysis:")
        for num,i in enumerate(chip_list):
            result = self.analyze.baseline_directory(folder_path[num], i[1])
            cell = ws.cell(row = 1 + next_space + num, column = 4)
            if (result == True):
                analysis_messages.append("Chip {} (Socket {}) PASSED!".format(i[1], i[0]))
                cell.value = "PASS"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
            if (result == False):
                analysis_messages.append("Chip {} (Socket {}) FAILED!".format(i[1], i[0]))
                cell.value = "FAIL"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
			
#Get data for the monitor tests
        for num,i in enumerate(chip_list):
            self.sbnd.save_monitor_data(folder_path[num], chip_index = i[0], chip_name = i[1])
#Analyze the monitor tests
        analysis_messages.append("Test Monitor Analysis:")
        for num,i in enumerate(chip_list):
            result = self.analyze2.monitor_directory(folder_path[num], i[1])
            cell = ws.cell(row = 1 + next_space + num, column = 7)
            if (result == True):
                analysis_messages.append("Chip {} (Socket {}) PASSED!".format(i[1], i[0]))
                cell.value = "PASS"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
            if (result == False):
                analysis_messages.append("Chip {} (Socket {}) FAILED!".format(i[1], i[0]))
                cell.value = "FAIL"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                
##Get data for the pulse calibration tests
#        for i in chip_list:
#            self.sbnd.save_pulse_data(folder_path[i[0]], chip_index = i[0], chip_name = i[1])
##Analyze the pulse calibration tests
#        for i in chip_list:
#            result = self.analyze2.pulse_directory(folder_path[i[0]], i[1])
#            if (result == True):
#                analysis_messages.append("Chip {} (Socket {}) PASSED!".format(i[1], i[0]))
#            if (result == False):
#                analysis_messages.append("Chip {} (Socket {}) FAILED!".format(i[1], i[0]))
                
#Get data for the channel alive tests
        self.sbnd.save_alive_data(folder_path, chip_list = chip_list)
#Analyze the channel alive tests
        analysis_messages.append("Channel Alive Analysis:")
        for num,i in enumerate(chip_list):
            result = self.analyze.alive_directory(folder_path[num], i[1])
            cell = ws.cell(row = 1 + next_space + num, column = 5)
            if (result == True):
                analysis_messages.append("Chip {} (Socket {}) PASSED!".format(i[1], i[0]))
                cell.value = "PASS"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.green, end_color = settings.green, fill_type = 'solid')
            if (result == False):
                analysis_messages.append("Chip {} (Socket {}) FAILED!".format(i[1], i[0]))
                cell.value = "FAIL"
                cell.font = Font(color = settings.white)
                cell.fill = PatternFill(start_color = settings.red, end_color = settings.red, fill_type = 'solid')
                
        #Tells the FPGA to turn off the ASICs
        self.sbnd.femb_config.femb.write_reg(12, 0xF)
        
        wb.save(filename = settings.root_path + settings.spreadsheet)

        #Print and save results
        screendisplay = sys.stdout
        for num,i in enumerate(chip_list):
            verdict_file = folder_path[num] + "Results.txt"
            sys.stdout = open(verdict_file, "a+")
            for j in analysis_messages:
                print (j)
            sys.stdout.close()
        
        sys.stdout = screendisplay
        for j in analysis_messages:
            print (j)
        print("Quad test complete!")

    def map_directory(self, folder):
        """
        This will put a config file in the main directory that 
        has the names of the other directories in it.  
        Say you do the Analysis days later, in a different Python 
        environment. 
        The first thing it does is open up this file to see 
        "oh, the baseline folder is called 'Test 1 and 2- Baseline and 
        RMS'", and then it'll move to that folder. 
        I did it this way because it's better than hardcoding in the
        folder structure, which is bound to change
        """
        to_save = dict()
        
        to_save["baseline_rms"] = settings.baseline_folder
        to_save["alive"] = settings.alive_folder
        to_save["pulse"] = settings.pulse_folder
        to_save["dac"] = settings.DAC_folder
        to_save["monitor"] = settings.monitor_folder
        to_save["data"] = settings.data
        
        with open(folder + 'directory_map.cfg', 'wb') as f:
            pickle.dump(to_save, f, pickle.HIGHEST_PROTOCOL)
            
#Loads or creates the spreadsheet that tracks every chip's results
    def get_results_sheet(self):
        if (not os.path.exists(settings.root_path + settings.spreadsheet)):
            wb = Workbook()
            wb.remove_sheet(wb.active)
            ws = wb.create_sheet()
            ws.title = "Results"
            ws.cell(row = 1, column = 1).value = "Datetime"
            ws.cell(row = 1, column = 2).value = "Chip"
            ws.cell(row = 1, column = 3).value = "Socket"
            ws.cell(row = 1, column = 4).value = "Baseline"
            ws.cell(row = 1, column = 5).value = "Alive"
            ws.cell(row = 1, column = 6).value = "DAC"
            ws.cell(row = 1, column = 7).value = "Monitor"
            for i in range(7):
                ws.cell(row = 1, column = 1 + i).alignment = Alignment(horizontal='center')
                ws.cell(row = 1, column = 1 + i).font = Font(bold=True)
        else:
            wb = load_workbook(filename = settings.root_path + settings.spreadsheet, read_only = False)
            ws = wb.get_sheet_by_name("Results")
            
        return wb
            
    def help_info(self):
        print ("Type in the function you want to call.")
        print ("Type 'test' to begin the Quad Board test for {} chips".format(settings.chip_num))
        print ("Type 'end' to exit.")

def main():
    '''
    Run set of FE tests.
    '''
    import json
    params = json.loads(open(sys.argv[1]).read())
    print (settings.path)
    if (not os.path.exists(settings.path)):
        os.makedirs(settings.path)
    food = main_class()
    food.check_setup()
    food.quad_test(**params)
            
if __name__ == "__main__":
    main()
